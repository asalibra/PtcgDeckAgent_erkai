#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Limitless TCG 多赛事环境综合统计
=================================
汇总 svi-jtg 环境下所有 300+ 人赛事的卡组数据，生成综合 Excel 报表。

用法:
    python3 limitless_meta_report.py
    python3 limitless_meta_report.py --output meta_report.xlsx
    python3 limitless_meta_report.py --min-players 300
"""

import json
import urllib.request
import time
import argparse
import sys
import os
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# 复用卡组中文名映射
from limitless_deck_analysis import DECK_CN, translate_deck_name

# ============================================================
# svi-jtg 环境 300+ 人赛事列表 (labs ID, 城市, 人数, 类型)
# ============================================================
SVI_JTG_TOURNAMENTS = [
    ("0025", "Atlanta",    2684, "regional"),
    ("0032", "Portland",   1688, "regional"),
    ("0028", "Milwaukee",  1657, "regional"),
    ("0026", "Monterrey",  1327, "regional"),
    ("0031", "Santiago",   1249, "regional"),
    ("0030", "Utrecht",    1241, "special"),
    ("0033", "Bologna",    1239, "special"),
    ("0027", "Sevilla",     817, "special"),
    ("0029", "Melbourne",   563, "regional"),
]

# 缺少 labs 数据的赛事:
# Malaysia Master Ball League (1017), Philippines MBL (515), Korean League S4 (308)


def fetch_json(url, retries=3):
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=20) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                if data.get("ok"):
                    return data["message"]
        except Exception as e:
            if attempt == retries - 1:
                print(f"    [WARN] Failed: {e}")
            time.sleep(0.5)
    return None


def fetch_all_deck_data(min_players=300):
    """拉取所有符合条件的赛事卡组数据。"""
    # 筛选赛事
    tournaments = [(tid, city, pl, tt) for tid, city, pl, tt in SVI_JTG_TOURNAMENTS if pl >= min_players]

    print(f"目标赛事: {len(tournaments)} 场 (>= {min_players} 人)")
    total_players = sum(t[2] for t in tournaments)
    print(f"总参赛人次: {total_players}")

    all_data = []  # [(tid, city, players_total, deck_data_list)]
    for tid, city, pl, tt in tournaments:
        print(f"\n  [{tid}] {city} ({pl}人, {tt})...")
        url = f"https://mew.limitlesstcg.com/labs/data/tcg/decks?tournamentId={tid}&division=MA"
        decks = fetch_json(url)
        if decks:
            print(f"    获取到 {len(decks)} 个卡组数据")
            all_data.append((tid, city, pl, decks))
        else:
            print(f"    获取失败!")
        time.sleep(0.3)

    return all_data


def aggregate_meta(all_data):
    """聚合所有赛事的卡组数据。"""
    # deck_identifier -> aggregated stats
    meta = defaultdict(lambda: {
        "name": "", "cn": "", "icons": "",
        "total_players": 0, "total_day2s": 0,
        "total_wins": 0, "total_losses": 0, "total_ties": 0,
        "tournaments": [],  # [(city, players, wins, losses, ties)]
    })

    for tid, city, t_players, decks in all_data:
        for d in decks:
            ident = d["identifier"]
            m = meta[ident]
            m["name"] = d["name"]
            m["cn"] = translate_deck_name(ident, d["name"])
            m["icons"] = d.get("icons", "")
            m["total_players"] += d["players"]
            m["total_day2s"] += d.get("day2s", 0)

            w = d["wins"]
            l = d["losses"]
            t = d["ties"]
            m["total_wins"] += w
            m["total_losses"] += l
            m["total_ties"] += t
            m["tournaments"].append((city, t_players, d["players"], w, l, t))

    return meta


# ============================================================
# Excel 样式
# ============================================================
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="Microsoft YaHei", size=10)
DATA_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
TITLE_FONT = Font(name="Microsoft YaHei", bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="2F5496")
SUMMARY_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
TIER_S = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # >=55% 绿
TIER_A = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # >=50% 浅绿
TIER_B = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # >=45% 黄
TIER_C = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # >=40% 橙
TIER_D = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # <40% 红


def wr_fill(wr):
    if wr >= 55: return TIER_S
    if wr >= 50: return TIER_A
    if wr >= 45: return TIER_B
    if wr >= 40: return TIER_C
    return TIER_D


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def style_cell(cell, is_left=False, win_rate=None):
    cell.font = DATA_FONT
    cell.alignment = LEFT_ALIGN if is_left else DATA_ALIGN
    cell.border = THIN_BORDER
    if win_rate is not None:
        cell.fill = wr_fill(win_rate)


# ============================================================
# Sheet 1: 环境总览
# ============================================================
def write_meta_overview(wb, meta, tournaments, min_players):
    ws = wb.create_sheet("环境总览")

    # 标题
    ws.merge_cells("A1:K1")
    ws["A1"].value = f"PTCG svi-jtg 环境综合统计 | {len(tournaments)} 场 300+ 人赛事 | 总参赛 {sum(t[2] for t in tournaments)} 人次"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # 赛事列表
    ws.merge_cells("A2:K2")
    event_names = ", ".join(f"{city}({pl})" for _, city, pl, _ in tournaments)
    ws["A2"].value = f"赛事: {event_names}"
    ws["A2"].font = Font(name="Microsoft YaHei", size=9, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 25

    # 表头
    headers = [
        "排名", "卡组(英)", "卡组(中)", "总使用人数",
        "总场次", "总胜", "总负", "总平",
        "综合胜率(%)", "使用占比(%)", "Day2人数"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    style_header(ws, 4, len(headers))

    # 排序: 按使用人数降序
    sorted_meta = sorted(meta.items(), key=lambda x: x[1]["total_players"], reverse=True)
    total_all_players = sum(m["total_players"] for _, m in sorted_meta)

    row = 5
    rank = 0
    for ident, m in sorted_meta:
        rank += 1
        total_games = m["total_wins"] + m["total_losses"] + m["total_ties"]
        wr = m["total_wins"] / (m["total_wins"] + m["total_losses"]) * 100 if (m["total_wins"] + m["total_losses"]) > 0 else 0
        usage = m["total_players"] / total_all_players * 100 if total_all_players > 0 else 0

        values = [
            rank, m["name"], m["cn"], m["total_players"],
            total_games, m["total_wins"], m["total_losses"], m["total_ties"],
            round(wr, 1), round(usage, 1), m["total_day2s"]
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            is_left = c in (2, 3)
            wr_val = wr if c == 9 else None
            style_cell(cell, is_left=is_left, win_rate=wr_val)
            if c == 9:
                cell.number_format = '0.0'
            if c == 10:
                cell.number_format = '0.0'
        row += 1

    # 合计行
    grand_games = sum(m["total_wins"] + m["total_losses"] + m["total_ties"] for _, m in sorted_meta)
    grand_w = sum(m["total_wins"] for _, m in sorted_meta)
    grand_l = sum(m["total_losses"] for _, m in sorted_meta)
    grand_t = sum(m["total_ties"] for _, m in sorted_meta)
    grand_wr = grand_w / (grand_w + grand_l) * 100 if (grand_w + grand_l) > 0 else 0

    ws.cell(row=row, column=1, value="")
    ws.cell(row=row, column=2, value="合计")
    ws.cell(row=row, column=3, value="")
    ws.cell(row=row, column=4, value=total_all_players)
    ws.cell(row=row, column=5, value=grand_games)
    ws.cell(row=row, column=6, value=grand_w)
    ws.cell(row=row, column=7, value=grand_l)
    ws.cell(row=row, column=8, value=grand_t)
    ws.cell(row=row, column=9, value=round(grand_wr, 1))
    ws.cell(row=row, column=10, value=100.0)
    ws.cell(row=row, column=11, value="")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name="Microsoft YaHei", bold=True, size=10)
        cell.fill = SUMMARY_FILL
        cell.alignment = DATA_ALIGN
        cell.border = THIN_BORDER

    # 列宽
    widths = [6, 25, 18, 12, 10, 8, 8, 8, 12, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return ws


# ============================================================
# Sheet 2: 各赛事分布
# ============================================================
def write_per_tournament(wb, meta, tournaments):
    ws = wb.create_sheet("各赛事卡组分布")

    ws.merge_cells("A1:H1")
    ws["A1"].value = "各赛事卡组使用人数与胜率分布"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 表头: 卡组 | 赛事1 胜率/人数 | 赛事2 ...
    headers = ["卡组(英)", "卡组(中)"]
    for _, city, pl, _ in tournaments:
        headers.append(f"{city}\n{pl}人")
    headers.append("平均胜率")

    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_header(ws, 3, len(headers))
    ws.row_dimensions[3].height = 35

    # 按总人数排序
    sorted_meta = sorted(meta.items(), key=lambda x: x[1]["total_players"], reverse=True)

    # 建立 tournament city -> index 映射
    city_order = [city for _, city, _, _ in tournaments]

    row = 4
    for ident, m in sorted_meta:
        ws.cell(row=row, column=1, value=m["name"])
        style_cell(ws.cell(row=row, column=1), is_left=True)
        ws.cell(row=row, column=2, value=m["cn"])
        style_cell(ws.cell(row=row, column=2), is_left=True)

        # 建立该卡组在各赛事的数据
        city_data = {}
        for city, t_pl, c_pl, w, l, t in m["tournaments"]:
            city_data[city] = (c_pl, w, l, t)

        total_w, total_l = 0, 0
        for i, city in enumerate(city_order):
            col = 3 + i
            if city in city_data:
                c_pl, w, l, t = city_data[city]
                total_games = w + l + t
                wr = w / (w + l) * 100 if (w + l) > 0 else 0
                total_w += w
                total_l += l
                val = f"{c_pl}人\n{wr:.0f}%"
                cell = ws.cell(row=row, column=col, value=val)
                style_cell(cell, win_rate=wr)
                cell.font = Font(name="Microsoft YaHei", size=9)
            else:
                cell = ws.cell(row=row, column=col, value="—")
                style_cell(cell)

        # 平均胜率
        avg_wr = total_w / (total_w + total_l) * 100 if (total_w + total_l) > 0 else 0
        cell = ws.cell(row=row, column=len(headers), value=round(avg_wr, 1))
        style_cell(cell, win_rate=avg_wr)
        cell.number_format = '0.0'

        row += 1

    # 列宽
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18
    for i in range(len(city_order)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 14
    ws.column_dimensions[get_column_letter(len(headers))].width = 12

    return ws


# ============================================================
# Sheet 3: Tier 分级
# ============================================================
def write_tier_sheet(wb, meta, min_usage=1):
    ws = wb.create_sheet("Tier分级")

    ws.merge_cells("A1:F1")
    ws["A1"].value = "环境 Tier 分级 (按综合胜率)"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["Tier", "卡组(英)", "卡组(中)", "使用人数", "综合胜率(%)", "使用占比(%)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_header(ws, 3, len(headers))

    sorted_meta = sorted(meta.items(), key=lambda x: x[1]["total_players"], reverse=True)
    total_players = sum(m["total_players"] for _, m in sorted_meta)

    # 分 tier
    tiers = {"S (>=55%)": [], "A (50-55%)": [], "B (45-50%)": [], "C (40-45%)": [], "D (<40%)": []}
    for ident, m in sorted_meta:
        if m["total_players"] < min_usage:
            continue
        w, l = m["total_wins"], m["total_losses"]
        wr = w / (w + l) * 100 if (w + l) > 0 else 0
        usage = m["total_players"] / total_players * 100
        entry = (m["name"], m["cn"], m["total_players"], round(wr, 1), round(usage, 1))

        if wr >= 55: tiers["S (>=55%)"].append(entry)
        elif wr >= 50: tiers["A (50-55%)"].append(entry)
        elif wr >= 45: tiers["B (45-50%)"].append(entry)
        elif wr >= 40: tiers["C (40-45%)"].append(entry)
        else: tiers["D (<40%)"].append(entry)

    tier_fills = {
        "S (>=55%)": TIER_S, "A (50-55%)": TIER_A,
        "B (45-50%)": TIER_B, "C (40-45%)": TIER_C, "D (<40%)": TIER_D
    }

    row = 4
    for tier_name, entries in tiers.items():
        if not entries:
            continue
        entries.sort(key=lambda x: x[3], reverse=True)
        for i, (name, cn, pl, wr, usage) in enumerate(entries):
            tier_label = tier_name if i == 0 else ""
            ws.cell(row=row, column=1, value=tier_label)
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=3, value=cn)
            ws.cell(row=row, column=4, value=pl)
            ws.cell(row=row, column=5, value=wr)
            ws.cell(row=row, column=6, value=usage)

            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=c)
                cell.font = DATA_FONT
                cell.alignment = LEFT_ALIGN if c in (1, 2, 3) else DATA_ALIGN
                cell.border = THIN_BORDER
                cell.fill = tier_fills[tier_name]
            row += 1

        # tier 分隔空行
        row += 1

    widths = [14, 25, 18, 12, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return ws


# ============================================================
# 主流程
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="Limitless TCG Meta Report")
    parser.add_argument("--min-players", type=int, default=300, help="Min players per tournament (default: 300)")
    parser.add_argument("--output", default=None, help="Output xlsx path")
    args = parser.parse_args()

    print("=" * 60)
    print("  PTCG svi-jtg 环境综合统计")
    print("=" * 60)

    print(f"\n[1/4] 拉取赛事数据...")
    all_data = fetch_all_deck_data(args.min_players)
    if not all_data:
        print("ERROR: 未获取到任何赛事数据")
        sys.exit(1)

    print(f"\n[2/4] 聚合卡组数据...")
    meta = aggregate_meta(all_data)
    print(f"  共 {len(meta)} 个不同卡组")

    total_w = sum(m["total_wins"] for m in meta.values())
    total_l = sum(m["total_losses"] for m in meta.values())
    print(f"  总对局: {total_w + total_l + sum(m['total_ties'] for m in meta.values())}")

    print(f"\n[3/4] 生成 Excel...")
    tournaments_used = [(tid, city, pl, tt) for tid, city, pl, tt in SVI_JTG_TOURNAMENTS if pl >= args.min_players]

    if args.output:
        output_path = args.output
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"meta_report_svi_jtg_{ts}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_meta_overview(wb, meta, tournaments_used, args.min_players)
    write_per_tournament(wb, meta, tournaments_used)
    write_tier_sheet(wb, meta, min_usage=5)

    wb.save(output_path)
    abs_path = os.path.abspath(output_path)

    print(f"\n[4/4] 完成!")
    print(f"{'=' * 60}")
    print(f"  赛事数: {len(all_data)}")
    print(f"  卡组种类: {len(meta)}")
    print(f"  文件: {abs_path}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()

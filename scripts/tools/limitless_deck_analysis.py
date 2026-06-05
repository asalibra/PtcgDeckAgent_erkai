#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Limitless TCG 赛事卡组对战分析工具
===================================
用途: 从 labs.limitlesstcg.com 抓取指定卡组的对战数据，
      生成中文 Excel 报表（含整体对战统计 + 各玩家详细表现）。

用法:
    python3 limitless_deck_analysis.py --tournament 0032 --deck tera-box --top 128
    python3 limitless_deck_analysis.py --tournament 0032 --deck gardevoir-ex-sv --top 64
"""

import json
import urllib.request
import time
import argparse
import sys
import os
from collections import defaultdict, OrderedDict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ============================================================
# 卡组中文名称映射 (deck identifier → 中文)
# 新增卡组时在此处添加即可
# ============================================================
DECK_CN = {
    # 主流卡组
    "tera-box":                 "钛晶Box",
    "raging-bolt-ogerpon":      "猛雷鼓厄诡椪",
    "gardevoir-ex-sv":          "沙奈朵",
    "dragapult":                "多龙巴鲁托",
    "dragapult-dusknoir":       "多龙黑夜魔灵",
    "flareon-noctowl":          "火伊布猫头夜鹰",
    "joltik-box":               "电蜘蛛Box",
    "n-zoroark":                "N索罗亚克",
    "gholdengo-dragapult":      "赛富豪多龙",
    "gholdengo-dudunsparce":    "赛富豪土龙节节",
    "roaring-moon-dudunsparce": "吼叫尾土龙节节",
    "terapagos-noctowl":        "太乐巴戈斯猫头夜鹰",
    "espathra-froslass":        "超能艳鸵雪妖女",
    "dragapult-charizard":      "多龙喷火龙",
    "dragapult-pidgeot":        "多龙大比鸟",
    "dragapult-pidgeot":        "多龙大比鸟",
    "toedscruel-ex":            "怒鹦哥",
    "toedscruel":               "怒鹦哥",
    "froslass-munkidori":       "雪妖女够赞狗",
    "mamoswine":                "象牙猪",
    "slowking":                 "呆呆王",
    "palafin":                  "海豚侠",
    "gholdengo":                "赛富豪",
    "charizard-dusknoir":       "喷火龙黑夜魔灵",
    "hydreigon-ex":             "三首恶龙",
    "hydreigon":                "三首恶龙",
    "milotic-farigiraf":        "美纳斯奇麒麟",
    "gholdengo":                "赛富豪",
    "charizard-ex":             "喷火龙",
    "charizard":                "喷火龙",
    "charizard-pidgeot":        "喷火龙大比鸟",
    "charizard-noctowl":        "喷火龙猫头夜鹰",
    "blissey-ex":               "幸福蛋",
    "blissey":                  "幸福蛋",
    "ceruledge-ex":             "灼翼龙",
    "ceruledge":                "灼翼龙",
    "archaludon-ex":            "铝钢桥龙",
    "archaludon":               "铝钢桥龙",
    "archaludon-zoroark":       "铝钢桥龙索罗亚克",
    "ancient-box":              "古代Box",
    "azumarill-ssp":            "玛力露丽",
    "azumarill":                "玛力露丽",
    "blaziken-ex-jtg":          "火焰鸡",
    "blaziken":                 "火焰鸡",
    "alolan-exeggutor-ex":      "阿罗拉椰蛋树",
    "aegislash-par":            "坚盾剑怪",
    "poison-terapagos":         "太晶毒",
    "poison-archaludon":        "毒铝钢桥龙",
    "hop-zacian":               "赫普藏玛然特",
    "roaring-moon":             "吼叫尾",
    "hydrapple-ogerpon":        "蜜集大蛇厄诡椪",
    "iono-bellibolt":           "奇树电肚蛙",
    "dragapult-iron-thorns":    "多龙铁荆棘",
    "okidogi":                  "够赞狗",
    "dragapult-froslass":       "多龙雪妖女",
    "future-hands":             "未来手",
    "flareon":                  "火伊布",
    "clefairy-ogerpon":         "皮皮厄诡椪",
    "noctowl":                  "猫头夜鹰",
}

# 宝可梦英文 → 中文（用于对战记录中的对手卡组名显示）
POKEMON_CN = {
    # 卡组名中出现的宝可梦
    "Dragapult":        "多龙巴鲁托",
    "Gardevoir":        "沙奈朵",
    "Raging Bolt":      "猛雷鼓",
    "Ogerpon":          "厄诡椪",
    "Flareon":          "火伊布",
    "Noctowl":          "猫头夜鹰",
    "Joltik":           "电蜘蛛",
    "Zoroark":          "索罗亚克",
    "Gholdengo":        "赛富豪",
    "Dudunsparce":      "土龙节节",
    "Roaring Moon":     "吼叫尾",
    "Terapagos":        "太乐巴戈斯",
    "Espathra":         "超能艳鸵",
    "Froslass":         "雪妖女",
    "Charizard":        "喷火龙",
    "Dusknoir":         "黑夜魔灵",
    "Pidgeot":          "大比鸟",
    "Toedscruel":       "怒鹦哥",
    "Munkidori":        "够赞狗",
    "Mamoswine":        "象牙猪",
    "Slowking":         "呆呆王",
    "Palafin":          "海豚侠",
    "Hydreigon":        "三首恶龙",
    "Milotic":          "美纳斯",
    "Farigiraf":        "奇麒麟",
    "Blissey":          "幸福蛋",
    "Ceruledge":        "灼翼龙",
    "Archaludon":       "铝钢桥龙",
    "Ancient":          "古代",
    "Azumarill":        "玛力露丽",
    "Blaziken":         "火焰鸡",
    "Exeggutor":        "椰蛋树",
    "Alolan":           "阿罗拉",
    "Aegislash":        "坚盾剑怪",
    "Poison":           "毒",
    "Hop":              "赫普",
    "Zacian":           "藏玛然特",
    "Hydrapple":        "蜜集大蛇",
    "Bellibolt":        "电肚蛙",
    "Iono":             "奇树",
    "Iron Thorns":      "铁荆棘",
    "Okidogi":          "够赞狗",
    "Future":           "未来",
    "Hands":            "手",
    "Clefairy":         "皮皮",
    "Lillie":           "莉莉艾",
    "Tera":             "太晶",
    "Box":              "Box",
    "N's":              "N的",
    "N":                "N",
    "Zoroark":          "索罗亚克",
    "Munkidori":        "够赞狗",
    "Espathra":         "超能艳鸵",
}


def translate_deck_name(identifier, fallback_name):
    """将卡组 identifier 翻译为中文，优先查映射表，否则用 fallback。"""
    if not identifier:
        # identifier 为 None 或空，尝试用 fallback_name 反查
        for _, v in DECK_CN.items():
            if v == fallback_name:
                return v
        # 逐词翻译 fallback_name
        parts = fallback_name.split() if fallback_name else []
        cn_parts = [POKEMON_CN.get(p, p) for p in parts]
        return "".join(cn_parts) if cn_parts else (fallback_name or "未知")
    key = identifier.lower().strip()
    if key in DECK_CN:
        return DECK_CN[key]
    # 尝试从英文名逐词翻译
    parts = fallback_name.split()
    cn_parts = []
    for p in parts:
        if p in POKEMON_CN:
            cn_parts.append(POKEMON_CN[p])
        else:
            cn_parts.append(p)
    return "".join(cn_parts) if cn_parts else fallback_name


# ============================================================
# 数据抓取
# ============================================================

def fetch_json(url, retries=3):
    """从 Limitless API 获取 JSON 数据。"""
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=20) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                if data.get("ok"):
                    return data["message"]
        except Exception as e:
            if attempt == retries - 1:
                print(f"  [WARN] Failed to fetch {url}: {e}")
            time.sleep(0.5)
    return None


def fetch_decks(tournament_id, division="MA"):
    """获取赛事所有卡组列表。"""
    url = f"https://mew.limitlesstcg.com/labs/data/tcg/decks?tournamentId={tournament_id}&division={division}"
    return fetch_json(url)


def fetch_players(tournament_id, deck_id, division="MA"):
    """获取指定卡组的所有玩家。"""
    url = f"https://mew.limitlesstcg.com/labs/data/tcg/players?tournamentId={tournament_id}&division={division}&deckId={deck_id}"
    return fetch_json(url)


def fetch_matches(tournament_id, player_id):
    """获取指定玩家的所有对战记录。"""
    url = f"https://mew.limitlesstcg.com/labs/data/tcg/matches?tournamentId={tournament_id}&playerId={player_id}"
    return fetch_json(url)


# ============================================================
# 数据聚合
# ============================================================

def aggregate_matchups(all_matches, deck_id, player_ids):
    """聚合对战数据，返回 (整体统计, 各玩家统计)。"""
    overall = defaultdict(lambda: {"W": 0, "L": 0, "T": 0})
    per_player = defaultdict(lambda: defaultdict(lambda: {"W": 0, "L": 0, "T": 0}))

    for m in all_matches:
        pid = None
        opp_deck_id = None
        opp_deck_name = None

        if m["p1_deck"] == deck_id and m["p1_id"] in player_ids:
            pid = m["p1_id"]
            opp_deck_id = m["p2_deck"]
            opp_deck_name = m["p2_deck_name"]
        elif m["p2_deck"] == deck_id and m["p2_id"] in player_ids:
            pid = m["p2_id"]
            opp_deck_id = m["p1_deck"]
            opp_deck_name = m["p1_deck_name"]
        else:
            continue

        winner = m["winner"]
        result = "T" if winner == 0 else ("W" if winner == pid else "L")

        overall[opp_deck_name]["_id"] = opp_deck_id
        overall[opp_deck_name][result] += 1
        per_player[pid][opp_deck_name]["_id"] = opp_deck_id
        per_player[pid][opp_deck_name][result] += 1

    return overall, per_player


# ============================================================
# Excel 生成
# ============================================================

# 样式常量
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="Microsoft YaHei", size=10)
DATA_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WIN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
LOSS_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
TITLE_FONT = Font(name="Microsoft YaHei", bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="4472C4")
SUMMARY_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")


def style_header_row(ws, row, col_count):
    """格式化表头行。"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def style_data_cell(cell, is_left=False, win_rate=None):
    """格式化数据单元格。"""
    cell.font = DATA_FONT
    cell.alignment = LEFT_ALIGN if is_left else DATA_ALIGN
    cell.border = THIN_BORDER
    if win_rate is not None:
        if win_rate >= 70:
            cell.fill = WIN_FILL
        elif win_rate < 40:
            cell.fill = LOSS_FILL


def write_overall_sheet(wb, deck_cn, overall, deck_players_map, tournament_id):
    """写入整体对战统计 sheet。"""
    ws = wb.create_sheet("整体对战统计")

    # 标题
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"{deck_cn} 整体对战各卡组胜率统计 (赛事 {tournament_id})"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 表头
    headers = ["对手卡组", "对手中文名", "对手人数", "场次", "胜", "负", "平", "胜率(%)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, len(headers))

    # 数据排序: 按场次降序
    sorted_items = sorted(overall.items(), key=lambda x: sum(v for k, v in x[1].items() if k in ("W", "L", "T")), reverse=True)

    row = 4
    total_w, total_l, total_t = 0, 0, 0
    for opp_name, rec in sorted_items:
        opp_id = rec.get("_id") or ""
        opp_cn = translate_deck_name(opp_id, opp_name)
        opp_count = deck_players_map.get(opp_name, deck_players_map.get(opp_id, "—"))
        w, l, t = rec["W"], rec["L"], rec["T"]
        total = w + l + t
        wr = w / (w + l) * 100 if (w + l) > 0 else 0

        total_w += w; total_l += l; total_t += t

        ws.cell(row=row, column=1, value=opp_name)
        ws.cell(row=row, column=2, value=opp_cn)
        ws.cell(row=row, column=3, value=opp_count)
        ws.cell(row=row, column=4, value=total)
        ws.cell(row=row, column=5, value=w)
        ws.cell(row=row, column=6, value=l)
        ws.cell(row=row, column=7, value=t)
        ws.cell(row=row, column=8, value=round(wr, 1))

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            is_left = col in (1, 2)
            wr_val = wr if col == 8 else None
            style_data_cell(cell, is_left=is_left, win_rate=wr_val)
            if col == 8:
                cell.number_format = '0.0'

        row += 1

    # 合计行
    grand = total_w + total_l + total_t
    grand_wr = total_w / (total_w + total_l) * 100 if (total_w + total_l) > 0 else 0
    ws.cell(row=row, column=1, value="合计")
    ws.cell(row=row, column=2, value="")
    ws.cell(row=row, column=3, value="")
    ws.cell(row=row, column=4, value=grand)
    ws.cell(row=row, column=5, value=total_w)
    ws.cell(row=row, column=6, value=total_l)
    ws.cell(row=row, column=7, value=total_t)
    ws.cell(row=row, column=8, value=round(grand_wr, 1))
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="Microsoft YaHei", bold=True, size=10)
        cell.fill = SUMMARY_FILL
        cell.alignment = DATA_ALIGN
        cell.border = THIN_BORDER

    # 列宽
    col_widths = [25, 20, 10, 8, 6, 6, 6, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return ws


def write_player_sheet(wb, deck_cn, per_player, player_info, tournament_id):
    """写入各玩家详细表现 sheet。"""
    ws = wb.create_sheet("各玩家详细表现")

    # 标题
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"{deck_cn} 各玩家（前128名）对战各卡组详细表现 (赛事 {tournament_id})"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["玩家名", "排名", "对手卡组", "对手中文名", "场次", "胜", "负", "平", "胜率(%)"]

    # 按排名排序
    sorted_players = sorted(per_player.keys(), key=lambda pid: player_info.get(pid, {}).get("placement", 9999))

    row = 3
    for pid in sorted_players:
        info = player_info[pid]
        pname = info["name"]
        prank = info["placement"]
        decks = per_player[pid]

        pw = sum(d["W"] for d in decks.values())
        pl = sum(d["L"] for d in decks.values())
        pt = sum(d["T"] for d in decks.values())
        pwr = pw / (pw + pl) * 100 if (pw + pl) > 0 else 0

        # 玩家标题行
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        player_title = ws.cell(row=row, column=1)
        player_title.value = f"{pname} (#{prank}) | {pw}W-{pl}L-{pt}T | 胜率 {pwr:.1f}%"
        player_title.font = SUBTITLE_FONT
        player_title.alignment = Alignment(horizontal="left", vertical="center")
        player_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        player_title.fill = player_fill
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER
            ws.cell(row=row, column=c).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        row += 1

        # 表头
        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=h)
        style_header_row(ws, row, len(headers))
        row += 1

        # 数据
        sorted_decks = sorted(decks.items(), key=lambda x: sum(v for k, v in x[1].items() if k in ("W", "L", "T")), reverse=True)
        for opp_name, rec in sorted_decks:
            opp_id = rec.get("_id") or ""
            opp_cn = translate_deck_name(opp_id, opp_name)
            w, l, t = rec["W"], rec["L"], rec["T"]
            total = w + l + t
            wr = w / (w + l) * 100 if (w + l) > 0 else 0

            ws.cell(row=row, column=1, value=pname)
            ws.cell(row=row, column=2, value=prank)
            ws.cell(row=row, column=3, value=opp_name)
            ws.cell(row=row, column=4, value=opp_cn)
            ws.cell(row=row, column=5, value=total)
            ws.cell(row=row, column=6, value=w)
            ws.cell(row=row, column=7, value=l)
            ws.cell(row=row, column=8, value=t)
            ws.cell(row=row, column=9, value=round(wr, 1))

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                is_left = col in (1, 3, 4)
                wr_val = wr if col == 9 else None
                style_data_cell(cell, is_left=is_left, win_rate=wr_val)
            row += 1

        row += 1  # 空行分隔

    # 列宽
    col_widths = [18, 8, 25, 20, 8, 6, 6, 6, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return ws


# ============================================================
# 主流程
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="Limitless TCG Deck Matchup Analyzer")
    parser.add_argument("--tournament", required=True, help="Tournament ID (e.g. 0032)")
    parser.add_argument("--deck", required=True, help="Deck identifier (e.g. tera-box)")
    parser.add_argument("--top", type=int, default=128, help="Top N placement filter (default: 128)")
    parser.add_argument("--division", default="MA", help="Division (default: MA)")
    parser.add_argument("--output", default=None, help="Output xlsx path (auto-generated if omitted)")
    args = parser.parse_args()

    tid = args.tournament
    deck_id = args.deck
    top_n = args.top
    division = args.division

    print(f"[1/5] 获取赛事 {tid} 卡组列表...")
    all_decks = fetch_decks(tid, division)
    if not all_decks:
        print("ERROR: 无法获取卡组数据")
        sys.exit(1)

    # 查找目标卡组
    target_deck = None
    deck_players_map = {}
    for d in all_decks:
        deck_players_map[d["name"]] = d["players"]
        deck_players_map[d["identifier"]] = d["players"]
        if d["identifier"] == deck_id:
            target_deck = d

    if not target_deck:
        print(f"ERROR: 未找到卡组 '{deck_id}'")
        print("可用卡组:", [d["identifier"] for d in all_decks[:20]])
        sys.exit(1)

    deck_cn = translate_deck_name(deck_id, target_deck["name"])
    print(f"  目标卡组: {target_deck['name']} ({deck_cn}) | 使用人数: {target_deck['players']}")

    print(f"[2/5] 获取 {deck_cn} 玩家列表...")
    players = fetch_players(tid, deck_id, division)
    if not players:
        print("ERROR: 无法获取玩家数据")
        sys.exit(1)

    top_players = [p for p in players if p["placement"] and p["placement"] <= top_n]
    print(f"  总玩家: {len(players)} | 前{top_n}名: {len(top_players)}")

    player_info = {}
    for p in top_players:
        player_info[p["tp_id"]] = p

    print(f"[3/5] 获取各玩家对战记录...")
    all_matches = []
    for i, p in enumerate(top_players):
        pid = p["tp_id"]
        pname = p["name"]
        matches = fetch_matches(tid, pid)
        if matches:
            all_matches.extend(matches)
            print(f"  [{i+1}/{len(top_players)}] {pname} (#{p['placement']}): {len(matches)} 场")
        else:
            print(f"  [{i+1}/{len(top_players)}] {pname}: 获取失败")
        time.sleep(0.3)

    print(f"[4/5] 聚合对战数据...")
    player_ids = set(player_info.keys())
    overall, per_player = aggregate_matchups(all_matches, deck_id, player_ids)

    total_games = sum(v for rec in overall.values() for k, v in rec.items() if k in ("W", "L", "T"))
    total_w = sum(rec["W"] for rec in overall.values())
    total_l = sum(rec["L"] for rec in overall.values())
    total_t = sum(rec["T"] for rec in overall.values())
    overall_wr = total_w / (total_w + total_l) * 100 if (total_w + total_l) > 0 else 0
    print(f"  总对局: {total_games} | {total_w}W-{total_l}L-{total_t}T | 胜率: {overall_wr:.1f}%")

    print(f"[5/5] 生成 Excel 报表...")
    if args.output:
        output_path = args.output
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"deck_analysis_{deck_id}_{tid}_{timestamp}.xlsx"

    wb = openpyxl.Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    write_overall_sheet(wb, deck_cn, overall, deck_players_map, tid)
    write_player_sheet(wb, deck_cn, per_player, player_info, tid)

    wb.save(output_path)
    abs_path = os.path.abspath(output_path)
    print(f"\n{'='*60}")
    print(f"  分析完成!")
    print(f"  卡组: {deck_cn} ({deck_id})")
    print(f"  赛事: {tid} | 前{top_n}名玩家: {len(top_players)}人")
    print(f"  总对局: {total_games} | 胜率: {overall_wr:.1f}%")
    print(f"  文件: {abs_path}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
